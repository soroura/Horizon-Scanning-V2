"""
Excel formatter — openpyxl workbook with triage colour coding and URL hyperlinks.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.module1_scanner.models import ScanItem
from src.module2_scorer.models import ScoreCard

# ── Triage row fill colours ───────────────────────────────────────────────────
_TRIAGE_FILLS = {
    "Act Now":       PatternFill("solid", fgColor="FF4444"),
    "Watch":         PatternFill("solid", fgColor="FF8C00"),
    "Monitor":       PatternFill("solid", fgColor="FFD700"),
    "For Awareness": PatternFill("solid", fgColor="90EE90"),
    "Low Signal":    PatternFill("solid", fgColor="D3D3D3"),
}

_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_COLUMNS = [
    ("Triage",          14),
    ("Title",           60),
    ("Source",          25),
    ("Published",       12),
    ("Horizon",          8),
    ("Preprint",         8),
    ("Domains",         25),
    ("Score",           7),
    ("Evidence (A)",    13),
    ("Impact (B)",      12),
    ("Insurance (C)",   13),
    ("Relevance (D)",   13),
    ("Annotation",      50),
    ("Suggested Action", 35),
    ("URL",             50),
]


def format_excel(
    scorecards: list[ScoreCard],
    items_by_id: dict[str, ScanItem],
    run_meta: dict,
) -> bytes:
    """
    Render Excel workbook and return bytes.
    One row per scored item, triage colour coding, URL hyperlinks.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horizon Scan Results"
    ws.freeze_panes = "A2"  # Freeze header row

    # ── Header row ─────────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_idx, sc in enumerate(scorecards, start=2):
        item = items_by_id.get(sc.item_id)
        if item is None:
            continue

        fill = _TRIAGE_FILLS.get(sc.triage_level, PatternFill())
        domains_str = ", ".join(item.domains)

        values = [
            f"{sc.triage_emoji} {sc.triage_level}",
            item.title,
            item.source_name,
            str(item.published_date),
            item.horizon_tier,
            "Yes" if item.is_preprint else "No",
            domains_str,
            round(sc.composite_score, 1),
            round(sc.evidence_strength, 1),
            round(sc.clinical_impact, 1),
            round(sc.insurance_readiness, 1),
            round(sc.domain_relevance, 1),
            sc.annotation,
            sc.suggested_action,
            item.url,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # URL column — make it a hyperlink
            if col_idx == 15 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single")

        ws.row_dimensions[row_idx].height = 45

    # ── Add run metadata sheet ─────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Run Info")
    ws_meta.append(["Run ID", run_meta.get("run_id", "")])
    ws_meta.append(["Profile", run_meta.get("profile_name", "")])
    ws_meta.append(["Date", run_meta.get("run_date", "")])
    ws_meta.append(["Sources scanned", run_meta.get("sources_count", 0)])
    ws_meta.append(["Items scored", len(scorecards)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
